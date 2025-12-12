"""
Excel Utilities - Cell Address Parsing and Table Reading

This module provides low-level Excel operations used by the extraction engine.
It handles the translation between Excel's A1-style cell addresses and Python's
row/column indices, as well as reading tabular data from worksheets.

ARCHITECTURAL CONTEXT:
----------------------
This is the lowest layer of the domain logic, providing primitive operations
that the engine module builds upon:

    engine.py (high-level extraction)
        │
        ▼
    excel.py (this module - cell/table operations)
        │
        ▼
    openpyxl (Excel file I/O)

This separation keeps the engine focused on business logic (what to extract)
while this module handles Excel mechanics (how to read cells).

CELL ADDRESSING:
----------------
Excel uses A1-style addresses where:
- Columns are letters: A, B, C, ..., Z, AA, AB, ..., AZ, BA, ...
- Rows are 1-based numbers: 1, 2, 3, ...

openpyxl uses 1-based (row, col) tuples internally:
- Row 1 is the first row
- Column 1 is column A, column 2 is column B, etc.

This module converts between these representations.

VALUE CONVERSION:
-----------------
Excel cells can contain various types: strings, numbers, dates, formulas, etc.
This module normalizes all values to strings for consistent downstream processing.

Special handling:
- Dates → YYYYMMDD format (no separators)
- Floats that are whole numbers → integers (removes ".0")
- None/empty → empty string
"""

import re
from datetime import datetime, date
from openpyxl.worksheet.worksheet import Worksheet


def parse_cell_address(address: str) -> tuple[int, int]:
    """
    Parse an A1-style cell address into (row, column) indices.

    Excel cell addresses combine a column letter sequence with a row number.
    This function converts that human-readable format into the numeric indices
    that openpyxl uses internally.

    Args:
        address: A1-style cell address (e.g., "A1", "B5", "AA10", "ZZ999").
                 Case-insensitive for the column letters.

    Returns:
        A tuple of (row, column) as 1-based integers.
        For example:
        - "A1" → (1, 1)
        - "B5" → (5, 2)
        - "AA10" → (10, 27)  # AA is the 27th column

    Raises:
        ValueError: If the address doesn't match the expected pattern.

    Column Calculation:
        Excel columns use base-26 numbering with letters A-Z.
        A=1, B=2, ..., Z=26, AA=27, AB=28, ..., AZ=52, BA=53, ...

        The formula is: col = Σ(letter_value × 26^position)
        For "AA": A(26^1) + A(26^0) = 26 + 1 = 27
    """
    # Match pattern: one or more letters followed by one or more digits
    match = re.match(r"([A-Za-z]+)(\d+)", address)
    if not match:
        raise ValueError(f"Invalid cell address: {address}")

    col_str, row_str = match.groups()

    # Convert column letters to number using base-26 arithmetic
    # Each letter position represents a power of 26
    col = 0
    for c in col_str.upper():
        # Shift existing value by one "digit" (×26) and add new letter value
        col = col * 26 + (ord(c) - ord("A") + 1)

    return int(row_str), col


def cell_to_string(cell) -> str:
    """
    Convert a cell's value to a string representation.

    This function normalizes Excel cell values to strings for consistent
    processing. Different value types are handled appropriately to produce
    output suitable for the ERP import format.

    Args:
        cell: An openpyxl Cell object (not just the value).
              We take the cell rather than cell.value to allow for
              potential future use of cell formatting information.

    Returns:
        A string representation of the cell's value:
        - None → "" (empty string)
        - datetime/date → "YYYYMMDD" (no separators, ERP-friendly)
        - float that equals an integer → "123" (not "123.0")
        - Everything else → str(value)

    Date Handling:
        ERP systems often expect dates in YYYYMMDD format without separators.
        This function automatically converts both datetime and date objects
        to this format.

    Number Handling:
        Excel often stores integers as floats (e.g., 5 becomes 5.0).
        To avoid "5.0" in output, we convert whole-number floats to integers
        before stringifying.
    """
    # Handle empty cells
    if cell.value is None:
        return ""

    # Handle datetime objects (Excel stores dates as datetime)
    # Convert to YYYYMMDD format for ERP compatibility
    if isinstance(cell.value, datetime):
        return cell.value.strftime("%Y%m%d")

    # Handle date objects (less common, but possible)
    if isinstance(cell.value, date):
        return cell.value.strftime("%Y%m%d")

    # Handle floats that are actually integers
    # Excel often stores "5" as 5.0, which would stringify as "5.0"
    if isinstance(cell.value, float) and cell.value == int(cell.value):
        return str(int(cell.value))

    # Default: convert to string
    return str(cell.value)


def read_table(sheet: Worksheet, start_address: str, end_value: str | None = None) -> list[dict[str, str]]:
    """
    Read a tabular region from an Excel worksheet.

    This function reads data organized as a table with:
    - A header row containing column names
    - Data rows below the header
    - Columns extending to the right until an empty header cell

    The result is a list of dictionaries where keys are column headers
    and values are the cell contents as strings.

    Args:
        sheet: An openpyxl Worksheet object to read from.

        start_address: A1-style address of the top-left cell of the header row.
                       Example: "A10" means headers start at row 10, column A.

        end_value: Optional sentinel value that marks end of data.
                   If the first column of a row equals this value, reading stops.
                   Example: "TOTAL" to stop before a totals row.
                   If None, reading stops at the first row with an empty first cell.

    Returns:
        A list of dictionaries, one per data row. Each dictionary maps
        column header names to cell values (as strings).

        Example return for a table with headers "SKU", "Qty", "Price":
        [
            {"SKU": "ABC123", "Qty": "10", "Price": "25.00"},
            {"SKU": "DEF456", "Qty": "5", "Price": "30.00"},
        ]

    Table Detection:
        - Headers: Read rightward from start_address until empty cell
        - Data: Read downward from row after headers until:
          - First cell is empty, OR
          - First cell equals end_value

    Edge Cases:
        - If no headers are found, returns empty list
        - If all columns are empty in header row, returns empty list
        - Headers with trailing spaces are preserved (not trimmed)
    """
    # Parse the starting cell address to get row and column indices
    start_row, start_col = parse_cell_address(start_address)

    # PHASE 1: Read header row
    # Scan rightward from the start column, collecting header names
    # Stop when we hit an empty cell (marks end of table columns)
    headers = []
    col = start_col
    while True:
        val = cell_to_string(sheet.cell(start_row, col))
        if not val.strip():
            # Empty header cell marks the end of columns
            break
        headers.append(val)
        col += 1

    # If no headers found, the table is empty or start_address is wrong
    if not headers:
        return []

    # PHASE 2: Read data rows
    # Scan downward from the row after headers
    # Stop when first cell is empty or matches end_value sentinel
    rows = []
    row_num = start_row + 1  # Data starts one row after headers

    while True:
        # Check the first cell to determine if we should continue
        first_cell = cell_to_string(sheet.cell(row_num, start_col))

        # Stop conditions:
        # 1. First cell is empty (or whitespace only)
        # 2. First cell matches the end sentinel value
        if not first_cell.strip() or (end_value and first_cell == end_value):
            break

        # Build a dictionary for this row: header_name → cell_value
        row_data = {
            headers[i]: cell_to_string(sheet.cell(row_num, start_col + i))
            for i in range(len(headers))
        }
        rows.append(row_data)
        row_num += 1

    return rows
