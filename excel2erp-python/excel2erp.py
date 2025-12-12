#!/usr/bin/env python3
"""
Excel2ERP - Python Implementation

Metadata-driven Excel to ERP conversion using Flask and openpyxl.
Demonstrates the algebraic approach: one configuration, many sources.
"""

import re
import io
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
from flask import Flask, request, send_file, render_template_string
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# Configuration Model
# =============================================================================

@dataclass
class SourceProperty:
    """Property extraction spec: name, cell/column locator, optional replacements."""
    name: str
    locator: str
    replacements: dict[str, str] = field(default_factory=dict)

    def convert(self, value: str) -> str:
        """Apply regex replacements to value."""
        result = value
        for pattern, replacement in self.replacements.items():
            result = re.sub(str(pattern), str(replacement), result)
        return result


@dataclass
class DetailConfig:
    """Detail table extraction: starting cell and column mappings."""
    locator: str
    properties: list[SourceProperty]
    end_value: str | None = None


@dataclass
class Source:
    """Source-specific extraction configuration."""
    name: str
    description: str
    sheet_index: int
    header: list[SourceProperty]
    detail: DetailConfig
    logo: str | None = None
    default_values: dict[str, str] = field(default_factory=dict)


@dataclass
class ResultProperty:
    """Output property with optional default and type."""
    name: str
    type: str = "text"
    prompt: str | None = None
    default_value: str | None = None

    def __post_init__(self):
        if self.prompt is None:
            self.prompt = self.name


@dataclass
class FileSpec:
    """Output file specification."""
    filename: str
    properties: list[ResultProperty]
    prolog: str = ""
    epilog: str = ""

    @property
    def default_values(self) -> dict[str, str]:
        return {p.name: p.default_value for p in self.properties if p.default_value}


@dataclass
class ResultConfig:
    """Result output configuration."""
    separator: str
    base_name: str
    header: FileSpec
    detail: FileSpec


@dataclass
class Config:
    """Full application configuration."""
    name: str
    description: str
    logo: str
    parameters: dict[str, str]
    sources: list[Source]
    result: ResultConfig

    def param(self, name: str) -> str:
        return self.parameters.get(name, name)


# =============================================================================
# Configuration Loader
# =============================================================================

def load_config(path: str) -> Config:
    """Load and parse YAML configuration."""
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)

    sources = []
    for src in data.get("sources", []):
        header_props = [
            SourceProperty(
                name=p["name"],
                locator=p["locator"],
                replacements=p.get("replacements", {})
            )
            for p in src.get("header", [])
        ]
        detail_props = [
            SourceProperty(
                name=p["name"],
                locator=p["locator"],
                replacements=p.get("replacements", {})
            )
            for p in src["detail"]["properties"]
        ]
        sources.append(Source(
            name=src["name"],
            description=src["description"],
            logo=src.get("logo"),
            sheet_index=src.get("sheetIndex", 0),
            header=header_props,
            detail=DetailConfig(
                locator=src["detail"]["locator"],
                properties=detail_props,
                end_value=src["detail"].get("endValue")
            ),
            default_values=src.get("defaultValues", {})
        ))

    def parse_file_spec(spec: dict) -> FileSpec:
        return FileSpec(
            filename=spec["filename"],
            prolog=spec.get("prolog", ""),
            epilog=spec.get("epilog", ""),
            properties=[
                ResultProperty(
                    name=p["name"],
                    type=p.get("type", "text"),
                    prompt=p.get("prompt"),
                    default_value=p.get("defaultValue")
                )
                for p in spec["properties"]
            ]
        )

    result = data["result"]
    return Config(
        name=data["name"],
        description=data["description"],
        logo=data.get("logo", ""),
        parameters=data.get("parameters", {}),
        sources=sources,
        result=ResultConfig(
            separator=result["separator"],
            base_name=result["baseName"],
            header=parse_file_spec(result["header"]),
            detail=parse_file_spec(result["detail"])
        )
    )


# =============================================================================
# Excel Utilities
# =============================================================================

def parse_cell_address(address: str) -> tuple[int, int]:
    """Parse A1-style address to (row, col) 1-based indices."""
    match = re.match(r"([A-Za-z]+)(\d+)", address)
    if not match:
        raise ValueError(f"Invalid cell address: {address}")
    col_str, row_str = match.groups()
    col = 0
    for c in col_str.upper():
        col = col * 26 + (ord(c) - ord("A") + 1)
    return int(row_str), col


def cell_to_string(cell) -> str:
    """Convert cell value to string."""
    from datetime import datetime, date
    if cell.value is None:
        return ""
    if isinstance(cell.value, datetime):
        return cell.value.strftime("%Y%m%d")
    if isinstance(cell.value, date):
        return cell.value.strftime("%Y%m%d")
    if isinstance(cell.value, float):
        if cell.value == int(cell.value):
            return str(int(cell.value))
        return str(cell.value)
    return str(cell.value)


def read_table(sheet: Worksheet, start_address: str, end_value: str | None = None) -> list[dict[str, str]]:
    """Read a table starting at given address, using first row as headers."""
    start_row, start_col = parse_cell_address(start_address)

    # Read header row to get column names
    headers = []
    col = start_col
    while True:
        val = cell_to_string(sheet.cell(start_row, col))
        if not val.strip():
            break
        headers.append(val)
        col += 1

    if not headers:
        return []

    # Read data rows
    rows = []
    row_num = start_row + 1
    while True:
        first_cell = cell_to_string(sheet.cell(row_num, start_col))
        if not first_cell.strip():
            break
        if end_value and first_cell == end_value:
            break

        row_data = {}
        for i, header in enumerate(headers):
            row_data[header] = cell_to_string(sheet.cell(row_num, start_col + i))
        rows.append(row_data)
        row_num += 1

    return rows


# =============================================================================
# Extraction Engine
# =============================================================================

def extract_header(sheet: Worksheet, source: Source, result_header: FileSpec) -> dict[str, str]:
    """Extract header fields from specific cells."""
    data = dict(result_header.default_values)
    data.update(source.default_values)

    for prop in source.header:
        row, col = parse_cell_address(prop.locator)
        value = cell_to_string(sheet.cell(row, col))
        data[prop.name] = prop.convert(value)

    return data


def extract_detail(sheet: Worksheet, source: Source) -> list[dict[str, str]]:
    """Extract detail rows from table."""
    raw_rows = read_table(sheet, source.detail.locator, source.detail.end_value)

    # Map column names to property names
    col_map = {p.locator: p for p in source.detail.properties}

    result = []
    for raw_row in raw_rows:
        row = {}
        for col_name, value in raw_row.items():
            if col_name in col_map:
                prop = col_map[col_name]
                row[prop.name] = prop.convert(value)
        result.append(row)

    return result


def missing_properties(source: Source, result_header: FileSpec) -> list[ResultProperty]:
    """Find header properties that need user input (not extracted, no default)."""
    extracted = {p.name for p in source.header}
    defaulted = set(source.default_values.keys()) | set(result_header.default_values.keys())
    return [p for p in result_header.properties if p.name not in extracted and p.name not in defaulted]


# =============================================================================
# Output Generation
# =============================================================================

def expand(template: str, props: dict[str, Any]) -> str:
    """Expand ${name} placeholders in template."""
    def replacer(match):
        key = match.group(1)
        return str(props.get(key, ""))
    return re.sub(r"\$\{(\w+)\}", replacer, template)


def normalize_date(value: str) -> str:
    """Remove non-digits from date string."""
    return re.sub(r"\D+", "", value)


def generate_content(spec: FileSpec, separator: str, records: list[dict[str, Any]]) -> str:
    """Generate file content from records."""
    lines = []
    if spec.prolog:
        lines.append(spec.prolog.rstrip())

    for idx, record in enumerate(records):
        values = []
        for prop in spec.properties:
            value = record.get(prop.name) or prop.default_value or ""
            if "${" in str(value):
                value = expand(str(value), {**record, "index": idx})
            values.append(str(value))
        lines.append(separator.join(values))

    if spec.epilog:
        lines.append(spec.epilog.rstrip())

    return "\n".join(lines)


def create_zip(header_content: str, detail_content: str, result: ResultConfig) -> bytes:
    """Create ZIP file with header and detail files."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(result.header.filename, header_content.encode("utf-8"))
        zf.writestr(result.detail.filename, detail_content.encode("utf-8"))
    return buffer.getvalue()


# =============================================================================
# Flask Application
# =============================================================================

INDEX_HTML = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ config.description }}</title>
    <script src="{{ config.param('htmx') }}"></script>
    <style>
        body { font-family: sans-serif; max-width: 800px; margin: 2em auto; padding: 0 1em; }
        .icon { height: 48px; vertical-align: middle; margin-right: 1em; }
        .source-logo { height: 32px; margin-bottom: 1em; }
        label { display: inline-block; min-width: 150px; }
        input, select { margin: 0.5em 0; padding: 0.3em; }
        input[type="submit"] { margin-top: 1em; padding: 0.5em 2em; }
        .form-row { margin: 0.5em 0; }
        hr { margin: 2em 0; }
        .error { color: red; }
    </style>
</head>
<body>
    <div>
        <img src="/assets/{{ config.logo }}" class="icon" alt="Logo">
        <h1 style="display:inline">{{ config.description }}</h1>
    </div>
    <br>
    <form action="/load" method="post" enctype="multipart/form-data">
        <div class="form-row">
            <label>{{ config.param('source') }}:</label>
            <select name="source" hx-get="/forms" hx-target="#source-form" hx-swap="innerHTML" required>
                <option value="">-- Seleccionar --</option>
                {% for source in config.sources %}
                <option value="{{ source.name }}">{{ source.description }}</option>
                {% endfor %}
            </select>
        </div>
        <div id="source-form"></div>
    </form>
    <hr>
    <a href="/close">Close</a>
</body>
</html>
"""

FORM_HTML = """
{% if src.logo %}
<img src="/assets/{{ src.logo }}" class="source-logo" alt="{{ src.description }}">
<hr>
{% endif %}
{% for prop in missing %}
<div class="form-row">
    <label>{{ prop.prompt }}:</label>
    <input type="{{ prop.type }}" name="{{ prop.name }}" required>
</div>
{% endfor %}
<div class="form-row">
    <label>{{ cfg.param('workbook') }}:</label>
    <input type="file" name="wbFile" accept=".xlsx" required>
</div>
<div class="form-row">
    <input type="submit" value="{{ cfg.param('submit') }}">
</div>
"""

ERROR_HTML = """
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Error</title></head>
<body>
    <div class="error" style="color: red; font-size: 150%;">
        <h1>Error</h1>
        <p>{{ message }}</p>
        <a href="javascript:history.back()">⇦ Volver</a>
    </div>
    {% if details %}
    <hr>
    <pre style="color: navy; font-size: 75%;">{{ details }}</pre>
    {% endif %}
</body>
</html>
"""


def create_app(config: Config, assets_dir: str) -> Flask:
    """Create and configure Flask application."""
    app = Flask(__name__)

    # Pre-compute missing properties per source
    source_map = {s.name: s for s in config.sources}
    missing_map = {s.name: missing_properties(s, config.result.header) for s in config.sources}

    @app.route("/")
    def index():
        return render_template_string(INDEX_HTML, config=config)

    @app.route("/assets/<path:filename>")
    def assets(filename):
        return send_file(Path(assets_dir) / filename)

    @app.route("/forms")
    def forms():
        source_name = request.args.get("source", "")
        if not source_name or source_name not in source_map:
            return ""
        source = source_map[source_name]
        missing = missing_map[source_name]
        return render_template_string(FORM_HTML, cfg=config, src=source, missing=missing)

    @app.route("/load", methods=["POST"])
    def load():
        try:
            source_name = request.form.get("source")
            if not source_name or source_name not in source_map:
                return render_template_string(ERROR_HTML, message="Fuente no seleccionada", details=None)

            source = source_map[source_name]
            file = request.files.get("wbFile")
            if not file or not file.filename:
                return render_template_string(ERROR_HTML, message="Archivo no proporcionado", details=None)

            if not file.filename.endswith(".xlsx"):
                return render_template_string(ERROR_HTML, message="Solo archivos .xlsx", details=None)

            # Load workbook
            wb = load_workbook(file, data_only=True)
            sheet = wb.worksheets[source.sheet_index]

            # Extract data
            header_data = extract_header(sheet, source, config.result.header)
            detail_data = extract_detail(sheet, source)

            # Add user-provided fields
            for prop in missing_map[source_name]:
                value = request.form.get(prop.name, "")
                if prop.type == "date":
                    value = normalize_date(value)
                header_data[prop.name] = value

            header_data["sourceName"] = source_name

            # Generate content
            header_content = generate_content(config.result.header, config.result.separator, [header_data])
            detail_content = generate_content(config.result.detail, config.result.separator, detail_data)

            # Create ZIP
            zip_bytes = create_zip(header_content, detail_content, config.result)
            zip_filename = expand(config.result.base_name, header_data) + ".zip"

            return send_file(
                io.BytesIO(zip_bytes),
                mimetype="application/zip",
                as_attachment=True,
                download_name=zip_filename
            )

        except Exception as e:
            import traceback
            return render_template_string(
                ERROR_HTML,
                message=config.param("extractionError"),
                details=traceback.format_exc()
            )

    @app.route("/close")
    def close():
        import os
        os._exit(0)

    return app


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(description="Excel2ERP Python Server")
    parser.add_argument("config", nargs="?", default="excel2erp.yaml", help="Config file path")
    parser.add_argument("--port", type=int, default=7070, help="Server port")
    parser.add_argument("--host", default="0.0.0.0", help="Server host")
    parser.add_argument("--basedir", default=".", help="Base directory")
    args = parser.parse_args()

    base_dir = Path(args.basedir).resolve()
    config_path = base_dir / args.config if not Path(args.config).is_absolute() else Path(args.config)
    assets_dir = base_dir / "assets"

    config = load_config(str(config_path))
    app = create_app(config, str(assets_dir))

    print(f"Server running at http://{args.host}:{args.port}")
    app.run(host=args.host, port=args.port, debug=False)


if __name__ == "__main__":
    main()
