"""
Tests for engine.py - extraction and output generation.

Includes parity tests comparing output against golden fixtures.
"""

import pytest
import zipfile
import io
from openpyxl import load_workbook

from config import load_config
from engine import (
    apply_replacements, extract_header, extract_detail,
    missing_properties, expand, normalize_date,
    generate_content, create_zip
)
from model import SourceProperty, ResultProperty, FileSpec


class TestApplyReplacements:
    def test_no_replacements(self):
        prop = SourceProperty(name="x", locator="A1", replacements={})
        assert apply_replacements("hello", prop) == "hello"

    def test_single_replacement(self):
        prop = SourceProperty(name="x", locator="A1", replacements={"^0+": ""})
        assert apply_replacements("00123", prop) == "123"

    def test_multiple_replacements(self):
        prop = SourceProperty(name="x", locator="A1", replacements={"-": "", " ": ""})
        assert apply_replacements("12-34 56", prop) == "123456"


class TestExpand:
    def test_simple_expansion(self):
        assert expand("${name}-${id}", {"name": "test", "id": "42"}) == "test-42"

    def test_missing_key_becomes_empty(self):
        assert expand("${name}-${missing}", {"name": "test"}) == "test-"

    def test_no_placeholders(self):
        assert expand("plain text", {}) == "plain text"


class TestNormalizeDate:
    def test_removes_separators(self):
        assert normalize_date("2024-01-15") == "20240115"
        assert normalize_date("2024/01/15") == "20240115"
        assert normalize_date("15.01.2024") == "15012024"

    def test_already_normalized(self):
        assert normalize_date("20240115") == "20240115"


class TestMissingProperties:
    def test_finds_properties_not_in_source(self, config_path):
        config = load_config(str(config_path))
        source = config.sources[0]

        missing = missing_properties(source, config.result.header)

        # Missing properties should not include those extracted or defaulted
        extracted_names = {p.name for p in source.header}
        defaulted_names = set(source.default_values.keys()) | set(config.result.header.default_values.keys())

        for prop in missing:
            assert prop.name not in extracted_names
            assert prop.name not in defaulted_names


class TestGenerateContent:
    def test_simple_output(self):
        spec = FileSpec(
            filename="test.txt",
            properties=[
                ResultProperty(name="a"),
                ResultProperty(name="b"),
            ]
        )
        records = [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]
        content = generate_content(spec, ";", records)

        assert content == "1;2\n3;4"

    def test_with_prolog_epilog(self):
        spec = FileSpec(
            filename="test.txt",
            prolog="HEADER",
            epilog="FOOTER",
            properties=[ResultProperty(name="x")]
        )
        records = [{"x": "data"}]
        content = generate_content(spec, ";", records)

        assert content == "HEADER\ndata\nFOOTER"

    def test_default_value_used(self):
        spec = FileSpec(
            filename="test.txt",
            properties=[
                ResultProperty(name="a"),
                ResultProperty(name="b", default_value="DEFAULT"),
            ]
        )
        records = [{"a": "1"}]
        content = generate_content(spec, ";", records)

        assert content == "1;DEFAULT"

    def test_index_expansion(self):
        spec = FileSpec(
            filename="test.txt",
            properties=[
                ResultProperty(name="line", default_value="${index}"),
            ]
        )
        records = [{}, {}, {}]
        content = generate_content(spec, ";", records)

        assert content == "0\n1\n2"


class TestCreateZip:
    def test_creates_valid_zip(self, config_path):
        config = load_config(str(config_path))

        header_content = "header;content"
        detail_content = "detail;content"

        zip_bytes = create_zip(header_content, detail_content, config.result)

        # Verify it's a valid ZIP
        with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
            names = zf.namelist()
            assert config.result.header.filename in names
            assert config.result.detail.filename in names

            # Verify content
            assert zf.read(config.result.header.filename).decode() == header_content
            assert zf.read(config.result.detail.filename).decode() == detail_content


class TestExtraction:
    """Integration tests using actual Excel fixtures."""

    def test_extract_header(self, config_path, sample_excel):
        config = load_config(str(config_path))
        source = config.sources[0]

        wb = load_workbook(sample_excel, data_only=True)
        sheet = wb.worksheets[source.sheet_index]

        header = extract_header(sheet, source, config.result.header)

        assert isinstance(header, dict)
        # Should have extracted values
        assert len(header) > 0

    def test_extract_detail(self, config_path, sample_excel):
        config = load_config(str(config_path))
        source = config.sources[0]

        wb = load_workbook(sample_excel, data_only=True)
        sheet = wb.worksheets[source.sheet_index]

        detail = extract_detail(sheet, source)

        assert isinstance(detail, list)
        assert len(detail) > 0
        assert all(isinstance(row, dict) for row in detail)


class TestParity:
    """
    Parity tests: verify output matches golden fixtures.
    All implementations must produce identical output.
    """

    def test_header_output_matches_expected(self, config_path, sample_excel, expected_dir):
        config = load_config(str(config_path))
        source = config.sources[0]

        wb = load_workbook(sample_excel, data_only=True)
        sheet = wb.worksheets[source.sheet_index]

        header_data = extract_header(sheet, source, config.result.header)

        # Add required fields that would come from user input
        missing = missing_properties(source, config.result.header)
        for prop in missing:
            if prop.name == "date":
                header_data[prop.name] = "20241215"
            else:
                header_data[prop.name] = ""

        header_data["sourceName"] = source.name

        content = generate_content(
            config.result.header,
            config.result.separator,
            [header_data]
        )

        expected_file = expected_dir / config.result.header.filename
        if expected_file.exists():
            expected = expected_file.read_text(encoding="utf-8").strip()
            assert content.strip() == expected

    def test_detail_output_matches_expected(self, config_path, sample_excel, expected_dir):
        config = load_config(str(config_path))
        source = config.sources[0]

        wb = load_workbook(sample_excel, data_only=True)
        sheet = wb.worksheets[source.sheet_index]

        detail_data = extract_detail(sheet, source)

        content = generate_content(
            config.result.detail,
            config.result.separator,
            detail_data
        )

        expected_file = expected_dir / config.result.detail.filename
        if expected_file.exists():
            expected = expected_file.read_text(encoding="utf-8").strip()
            assert content.strip() == expected
